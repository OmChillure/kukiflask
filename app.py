import os
import re
import google.generativeai as genai
from PyPDF2 import PdfReader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_google_genai import GoogleGenerativeAIEmbeddings
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_community.vectorstores import FAISS
from langchain.chains.question_answering import load_qa_chain
from langchain.prompts import PromptTemplate
from dotenv import load_dotenv
from docx import Document
import csv
from openpyxl import load_workbook
import streamlit as st

# Load environment variables
load_dotenv()
api = os.getenv("GOOGLE_API_KEY")
genai.configure(api_key=api)

# Function to extract text from docx file
def extract_text_from_docx(docx_file):
    document = Document(docx_file)
    return '\n'.join([paragraph.text for paragraph in document.paragraphs])

# Function to extract text from xlsx file
def extract_text_from_xlsx(xlsx_file):
    workbook = load_workbook(xlsx_file, data_only=True)
    text = ""
    for sheet in workbook:
        for row in sheet.iter_rows(values_only=True):
            row_text = ", ".join([str(cell) for cell in row if cell is not None])
            text += row_text + "\n"
    return text

# Function to extract text from csv file
def extract_text_from_csv(csv_file):
    content = []
    reader = csv.reader(csv_file)
    for row in reader:
        content.append(', '.join(row))
    return '\n'.join(content)

# Function to get text from documents with metadata
def get_text_from_documents(docs):
    texts_with_metadata = []
    for doc in docs:
        if doc.name.endswith('.pdf'):
            pdf_reader = PdfReader(doc)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            texts_with_metadata.append((text, {"source": doc.name}))
        elif doc.name.endswith('.docx'):
            text = extract_text_from_docx(doc) + "\n"
            texts_with_metadata.append((text, {"source": doc.name}))
        elif doc.name.endswith('.csv'):
            text = extract_text_from_csv(doc) + "\n"
            texts_with_metadata.append((text, {"source": doc.name}))
        elif doc.name.endswith('.xlsx'):
            text = extract_text_from_xlsx(doc) + "\n"
            texts_with_metadata.append((text, {"source": doc.name}))
    return texts_with_metadata

# Function to split text into chunks with metadata
def get_text_chunks_with_metadata(texts_with_metadata):
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=10000, chunk_overlap=1000)
    chunks_with_metadata = []
    for text, metadata in texts_with_metadata:
        chunks = text_splitter.split_text(text)
        chunks_with_metadata.extend([(chunk, metadata) for chunk in chunks])
    return chunks_with_metadata

# Function to get vectors with metadata
def get_vectors_with_metadata(chunks_with_metadata):
    embeddings = GoogleGenerativeAIEmbeddings(model='models/embedding-001')
    vector_store = FAISS.from_texts([chunk for chunk, metadata in chunks_with_metadata], embedding=embeddings, metadatas=[metadata for chunk, metadata in chunks_with_metadata])
    vector_store.save_local("faiss_index")

# Function to create a conversation chain
def convo_chain():
    prompt_template = """
        Please respond to the following questions. Answer all the questions based on the uploaded files also give information about the pdf and page number from which data is taken.

        1. Details of total energy consumption (in Joules or multiples) and energy intensity:

        | Parameter                          | FY 2024 (Current Financial Year) | FY __ (Previous Financial Year) |
        | Total electricity consumption (A)  |                                  |                                   |
        | Total fuel consumption (B)         |                                  |                                   |
        | Energy consumption through other sources (C) |                      |                                   |
        | Total energy consumption (A+B+C)   |                                  |                                   |
        | Energy intensity per rupee of turnover (Total energy consumption/ turnover in rupees) | |                                   |
        | Energy intensity (optional) – the relevant metric may be selected by the entity | |                                   |
        | Independent assessment carried out by an external agency? (Y/N) If yes, name of the external agency. | |                                   |

        3. Provide details of the following disclosures related to water:

        | Parameter                          | FY __ (Current Financial Year) | FY __ (Previous Financial Year) |
        | Water withdrawal by source (in kilolitres) |                          |                                   |
        | (i) Surface water                  |                                  |                                   |
        | (ii) Groundwater                   |                                  |                                   |
        | (iii) Third party water            |                                  |                                   |
        | (iv) Seawater / desalinated water  |                                  |                                   |
        | (v) Others                         |                                  |                                   |
        | Total volume of water withdrawal (i + ii + iii + iv + v) |            |                                   |
        | Total volume of water consumption  |                                  |                                   |
        | Water intensity per rupee of turnover (Water consumed / turnover) |   |                                   |
        | Water intensity (optional) – the relevant metric may be selected by the entity | |                                   |
        | Independent assessment carried out by an external agency? (Y/N) If yes, name of the external agency. | |                                   |

        4. Has the entity implemented a mechanism for Zero Liquid Discharge? If yes, provide details of its coverage and implementation.

        5. Please provide details of air emissions (other than GHG emissions) by the entity:

        | Parameter                          | Unit                             | FY __ (Current Financial Year) | FY __ (Previous Financial Year) |       
        | NOx                                |                                  |                                  |                                   |
        | SOx                                |                                  |                                  |                                   |
        | Particulate matter (PM)            |                                  |                                  |                                   |
        | Persistent organic pollutants (POP)|                                  |                                  |                                   |
        | Volatile organic compounds (VOC)   |                                  |                                  |                                   |
        | Hazardous air pollutants (HAP)     |                                  |                                  |                                   |
        | Others – please specify            |                                  |                                  |                                   |
        | Independent assessment carried out by an external agency? (Y/N) If yes, name of the external agency. |                                   |                                   |

        6. Provide details of greenhouse gas emissions (Scope 1 and Scope 2 emissions) & its intensity:

        | Parameter                          | Unit                             | FY __ (Current Financial Year) | FY __ (Previous Financial Year) |
        | Total Scope 1 emissions            | Metric tonnes of CO2 equivalent  |                                  |                                   |
        | Total Scope 2 emissions            | Metric tonnes of CO2 equivalent  |                                  |                                   |
        | Total Scope 1 and Scope 2 emissions per rupee of turnover |           |                                  |                                   |
        | Total Scope 1 and Scope 2 emission intensity (optional) – the relevant metric may be selected by the entity |                                   |                                   |
        | Independent assessment carried out by an external agency? (Y/N) If yes, name of the external agency. |                                   |                                   |

        Continue with similar formatting for the rest of the questions in the document.
        
        Text: {context}
    """
    prompt = PromptTemplate(template=prompt_template, input_variables=['context'])
    model = ChatGoogleGenerativeAI(model='gemini-pro', temperature=0.3)
    chain = load_qa_chain(model, chain_type='stuff', prompt=prompt)
    return chain

# Function to get user input with citations
def user_input(user_question):
    embeddings = GoogleGenerativeAIEmbeddings(model='models/embedding-001')
    new_db = FAISS.load_local("faiss_index", embeddings, allow_dangerous_deserialization=True)
    retriever = new_db.as_retriever()
    docs = retriever.invoke(user_question)
    chain = convo_chain()
    try:
        response = chain({"input_documents": docs, "question": user_question}, return_only_outputs=True)
        citations = [f"Document {i+1}: {doc.metadata['source']}" for i, doc in enumerate(docs)]
        return response['output_text'] + "\n\n" + "Citations:\n" + "\n".join(citations)
    except Exception as e:
        return str(e)

# Function to remove table formatting characters
def clean_for_insights(qa_results):
    cleaned_text = re.sub(r'\|', '', qa_results)
    cleaned_text = re.sub(r'\s*\n\s*', ' ', cleaned_text).strip()
    cleaned_text = re.sub(r'\s{2,}', ' ', cleaned_text)
    return cleaned_text

# Function to generate insights
def generate_insights(llm_response):
    overall_insight_question = "Provide a brief, high-level insight based on the following information:"
    overall_insight = user_input(overall_insight_question + llm_response)
    cleaned_text_for_insights = clean_for_insights(overall_insight)
    
    return {"Overall Insight": cleaned_text_for_insights}

# Function to clean text and format it as an HTML table
def format_as_html_table(text, citations):
    sections = re.split(r'\n\s*\n', text)
    html_output = ""
    citation_index = 0
    
    for section in sections:
        lines = section.strip().split('\n')
        html_table = "<table border='1'>"
        
        for line in lines:
            if line.startswith('| Parameter'):
                html_table += "<thead><tr>"
                cells = line.split('|')[1:-1]
                for cell in cells:
                    html_table += f"<th>{cell.strip()}</th>"
                html_table += "</tr></thead>"
            else:
                html_table += "<tr>"
                cells = line.split('|')[1:-1]
                for cell in cells:
                    html_table += f"<td>{cell.strip()}</td>"
                html_table += "</tr>"
        
        html_table += "</table>"
        
        # Append citation for the section
        if citation_index < len(citations):
            html_table += f"<p><em>{citations[citation_index]}</em></p>"
            citation_index += 1
        
        html_output += html_table
    
    return html_output

# Streamlit application
def main():
    st.title("Document Processing and Insights Generation")

    # Upload files
    uploaded_files = st.file_uploader("Upload documents", type=['pdf', 'docx', 'csv', 'xlsx'], accept_multiple_files=True)

    if st.button("Process Documents"):
        if not uploaded_files:
            st.warning("Please upload at least one document.")
        else:
            texts_with_metadata = get_text_from_documents(uploaded_files)
            chunks_with_metadata = get_text_chunks_with_metadata(texts_with_metadata)
            get_vectors_with_metadata(chunks_with_metadata)
            
            # Generate formatted table
            raw_text = " ".join([text for text, metadata in texts_with_metadata])
            qa_results = user_input(raw_text)
            citations = [f"Document {i+1}: {doc.name}" for i, doc in enumerate(uploaded_files)]
            formatted_table = format_as_html_table(qa_results, citations)

            # Generate insights
            cleaned_text_for_insights = clean_for_insights(qa_results)
            insights = generate_insights(cleaned_text_for_insights)

            # Display results
            st.markdown("### Formatted Table")
            st.markdown(formatted_table, unsafe_allow_html=True)
            
            st.markdown("### Insights")
            st.json(insights)

if __name__ == "__main__":
    main()
